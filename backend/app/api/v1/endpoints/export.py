"""
Routes pour l'export de données (Excel, CSV)
"""
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional, List
from datetime import datetime
import io
import csv

from app.core.database import get_db
from app.core.deps import get_current_active_user
from app.models.essai import Essai, TypeEssai, StatutEssai
from app.models.user import User

router = APIRouter()


def generate_csv_content(essais: List[Essai]) -> str:
    """Génère le contenu CSV des essais"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
    
    # En-têtes
    writer.writerow([
        'ID', 'Numéro Essai', 'Type', 'Statut', 'Projet', 'Échantillon',
        'Date Essai', 'Date Réception', 'Opérateur', 'Observations',
        'Résultats (JSON)'
    ])
    
    # Données
    for essai in essais:
        writer.writerow([
            essai.id,
            essai.numero_essai,
            essai.type_essai.value if essai.type_essai else '',
            essai.statut.value if essai.statut else '',
            essai.projet or '',
            essai.echantillon or '',
            essai.date_essai.isoformat() if essai.date_essai else '',
            essai.date_reception.isoformat() if essai.date_reception else '',
            essai.operateur.full_name if essai.operateur else essai.operateur.username if essai.operateur else '',
            essai.observations or '',
            str(essai.resultats) if essai.resultats else ''
        ])
    
    return output.getvalue()


def generate_excel_content(essais: List[Essai]) -> bytes:
    """Génère le contenu Excel des essais"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Essais"
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = [
        'ID', 'Numéro Essai', 'Type', 'Statut', 'Projet', 'Échantillon',
        'Date Essai', 'Date Réception', 'Opérateur', 'Observations'
    ]
    ws.append(headers)
    
    # Appliquer le style à l'en-tête
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Données
    for essai in essais:
        row = [
            essai.id,
            essai.numero_essai,
            essai.type_essai.value if essai.type_essai else '',
            essai.statut.value if essai.statut else '',
            essai.projet or '',
            essai.echantillon or '',
            essai.date_essai.isoformat() if essai.date_essai else '',
            essai.date_reception.isoformat() if essai.date_reception else '',
            essai.operateur.full_name if essai.operateur else essai.operateur.username if essai.operateur else '',
            essai.observations or ''
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/essais/csv")
async def export_essais_csv(
    type_essai: Optional[TypeEssai] = None,
    statut: Optional[StatutEssai] = None,
    search: Optional[str] = None,
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Exporte les essais en format CSV"""
    query = db.query(Essai)
    
    # Appliquer les filtres
    if type_essai:
        query = query.filter(Essai.type_essai == type_essai)
    if statut:
        query = query.filter(Essai.statut == statut)
    if search:
        query = query.filter(
            or_(
                Essai.numero_essai.ilike(f"%{search}%"),
                Essai.projet.ilike(f"%{search}%"),
                Essai.echantillon.ilike(f"%{search}%")
            )
        )
    if date_debut:
        query = query.filter(Essai.date_essai >= date_debut)
    if date_fin:
        query = query.filter(Essai.date_essai <= date_fin)
    
    essais = query.order_by(Essai.created_at.desc()).all()
    
    csv_content = generate_csv_content(essais)
    
    filename = f"essais_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.BytesIO(csv_content.encode('utf-8-sig')),  # utf-8-sig pour Excel
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/essais/excel")
async def export_essais_excel(
    type_essai: Optional[TypeEssai] = None,
    statut: Optional[StatutEssai] = None,
    search: Optional[str] = None,
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Exporte les essais en format Excel"""
    query = db.query(Essai)
    
    # Appliquer les filtres
    if type_essai:
        query = query.filter(Essai.type_essai == type_essai)
    if statut:
        query = query.filter(Essai.statut == statut)
    if search:
        query = query.filter(
            or_(
                Essai.numero_essai.ilike(f"%{search}%"),
                Essai.projet.ilike(f"%{search}%"),
                Essai.echantillon.ilike(f"%{search}%")
            )
        )
    if date_debut:
        query = query.filter(Essai.date_essai >= date_debut)
    if date_fin:
        query = query.filter(Essai.date_essai <= date_fin)
    
    essais = query.order_by(Essai.created_at.desc()).all()
    
    excel_content = generate_excel_content(essais)
    
    filename = f"essais_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

